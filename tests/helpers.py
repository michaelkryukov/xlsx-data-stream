import io
from collections.abc import AsyncIterable, Iterable
from datetime import datetime
from typing import TypedDict

from openpyxl import Workbook, load_workbook


class MovieDict(TypedDict):
    title: str
    rating: float
    release_date: datetime


MOVIES: list[MovieDict] = [
    {
        "title": "As Above, So Below",
        "rating": 0.63,
        "release_date": datetime.fromisoformat("2014-08-29 08:00:00+00:00"),
    },
    {
        "title": "The Machine Girl",
        "rating": 0.59,
        "release_date": datetime.fromisoformat("2008-08-02 09:00:00+00:00"),
    },
    {
        "title": "This Means War",
        "rating": 0.63,
        "release_date": datetime.fromisoformat("2012-02-08T10:00:00+00:00"),
    },
]


def get_movies():
    yield from MOVIES


async def get_movies_async():
    for movie in MOVIES:
        yield movie


async def list_to_async_iterable(values):
    for value in values:
        yield value


def join_chunks(chunks: Iterable[bytes]):
    return b"".join(list(chunks))


async def join_chunks_async(chunks: AsyncIterable[bytes]):
    return b"".join([chunk async for chunk in chunks])


def load_workbook_from_bytes(content: bytes) -> Workbook:
    with io.BytesIO(content) as fh:
        return load_workbook(fh)
