import pytest
from openpyxl import Workbook

from xlsx_data_stream import (
    ColumnDescription,
    get_xlsx_file_stream,
    get_xlsx_file_stream_async,
)

from .helpers import (
    MOVIES,
    get_movies,
    get_movies_async,
    join_chunks,
    join_chunks_async,
    load_workbook_from_bytes,
)

COLUMNS_DESCRIPTION = [
    ColumnDescription(
        field="title",
        title="Title",
        number_format="General",
    ),
    ColumnDescription(
        field="rating",
        title="Rating",
        number_format="0.00",
    ),
    ColumnDescription(
        field="release_date",
        title="Release date",
        number_format="dd.mm.yyyy hh:mm",
    ),
]


def _check_workbook(wb: Workbook):
    assert wb.active
    assert wb.active.title == "Sheet"
    assert wb.sheetnames == ["Sheet"]

    sheet = wb["Sheet"]
    assert sheet.freeze_panes == "A2"
    assert sheet.cell(1, 1).value == "Title"
    assert sheet.cell(1, 2).value == "Rating"
    assert sheet.cell(1, 3).value == "Release date"

    for row_number, movie in enumerate(MOVIES, 2):
        assert sheet.cell(row_number, 1).value == movie["title"]
        assert sheet.cell(row_number, 2).value == movie["rating"]

        naive_release_date = movie["release_date"].replace(tzinfo=None)
        assert sheet.cell(row_number, 3).value == naive_release_date


def test_simple():
    chunks = get_xlsx_file_stream(
        COLUMNS_DESCRIPTION,
        get_movies(),
    )

    content = join_chunks(chunks)

    _check_workbook(load_workbook_from_bytes(content))


@pytest.mark.asyncio
async def test_simple_async():
    chunks = get_xlsx_file_stream_async(
        COLUMNS_DESCRIPTION,
        get_movies_async(),
    )

    content = await join_chunks_async(chunks)

    _check_workbook(load_workbook_from_bytes(content))
