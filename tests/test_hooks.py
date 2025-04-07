import pytest
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from xlsx_data_stream import (
    ColumnDescription,
    get_xlsx_file_stream,
    get_xlsx_file_stream_async,
)

from .helpers import (
    get_movies,
    get_movies_async,
    join_chunks,
    join_chunks_async,
    load_workbook_from_bytes,
)


def _modify_template_worksheet(ws: Worksheet):
    ws.title = "Movies"
    ws.cell(1, 1).border = Border(bottom=Side("medium"))


def _modify_template_workbook(wb: Workbook):
    ws: Worksheet = wb.create_sheet("Information")
    ws.cell(1, 1).value = "Wow!"


def _check_workbook(wb: Workbook):
    assert wb.active
    assert wb.active.title == "Movies"
    assert wb.sheetnames == ["Movies", "Information"]
    assert wb["Movies"].freeze_panes == "A2"
    assert wb["Movies"].cell(1, 1).border.bottom.style == "medium"
    assert wb["Information"].freeze_panes is None
    assert wb["Information"].cell(1, 1).value == "Wow!"


COLUMNS_DESCRIPTION = [
    ColumnDescription(
        field="title",
        title="Title",
        number_format="General",
    ),
    ColumnDescription(
        field="rating",
        title="Rating",
        number_format="0.00",
    ),
]


def test_hooks():
    chunks = get_xlsx_file_stream(
        COLUMNS_DESCRIPTION,
        get_movies(),
        modify_template_worksheet=_modify_template_worksheet,
        modify_template_workbook=_modify_template_workbook,
    )

    content = join_chunks(chunks)

    _check_workbook(load_workbook_from_bytes(content))


@pytest.mark.asyncio
async def test_hooks_async():
    chunks = get_xlsx_file_stream_async(
        COLUMNS_DESCRIPTION,
        get_movies_async(),
        modify_template_worksheet=_modify_template_worksheet,
        modify_template_workbook=_modify_template_workbook,
    )

    content = await join_chunks_async(chunks)

    _check_workbook(load_workbook_from_bytes(content))
