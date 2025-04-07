from datetime import datetime, time, timedelta, timezone

import pytest
from openpyxl.workbook.workbook import Workbook

from xlsx_data_stream import (
    ColumnDescription,
    get_xlsx_file_stream,
)
from xlsx_data_stream.astream import get_xlsx_file_stream_async

from .helpers import (
    join_chunks,
    join_chunks_async,
    list_to_async_iterable,
    load_workbook_from_bytes,
)


def _check_workbook(wb: Workbook):
    assert wb.active

    cell = wb.active.cell(2, 1)
    assert cell.value == datetime(2020, 1, 1, 12, 0)
    assert cell.number_format == COLUMNS_DESCRIPTION[0].number_format

    cell = wb.active.cell(2, 2)
    assert cell.value == datetime(2020, 1, 1, 3, 0)
    assert cell.number_format == COLUMNS_DESCRIPTION[1].number_format

    cell = wb.active.cell(2, 3)
    assert cell.value == time(18, 0)
    assert cell.number_format == COLUMNS_DESCRIPTION[2].number_format

    cell = wb.active.cell(2, 4)
    assert cell.value == time(1, 0)
    assert cell.number_format == COLUMNS_DESCRIPTION[3].number_format


COLUMNS_DESCRIPTION = [
    ColumnDescription(
        field="datetime",
        number_format="dd.mm.yyyy hh:mm",
    ),
    ColumnDescription(
        field="date",
        number_format="dd.mm.yyyy",
    ),
    ColumnDescription(
        field="time",
        number_format="hh:mm",
    ),
    ColumnDescription(
        field="timedelta",
        number_format="hh:mm",
    ),
]

DOCUMENTS = [
    {
        "datetime": datetime(2020, 1, 1, 9, tzinfo=timezone.utc),
        "date": datetime(2020, 1, 1, 12, tzinfo=timezone.utc).date(),
        "time": datetime(2020, 1, 1, 15, 0, tzinfo=timezone.utc).time(),
        "timedelta": timedelta(hours=1),
    },
]


def test_moments():
    chunks = get_xlsx_file_stream(
        COLUMNS_DESCRIPTION,
        DOCUMENTS,
        display_timezone=timezone(timedelta(hours=3), "Europe/Moscow"),
    )

    content = join_chunks(chunks)

    _check_workbook(load_workbook_from_bytes(content))


@pytest.mark.asyncio
async def test_moments_async():
    chunks = get_xlsx_file_stream_async(
        COLUMNS_DESCRIPTION,
        list_to_async_iterable(DOCUMENTS),
        display_timezone=timezone(timedelta(hours=3), "Europe/Moscow"),
    )

    content = await join_chunks_async(chunks)

    _check_workbook(load_workbook_from_bytes(content))
